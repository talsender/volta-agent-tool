"""
CRM Knowledge Analysis - Volta Solar
Processes Residential - Last view used.xlsx (20,000 rows)
Outputs: crm_knowledge_analysis_output.xlsx + crm_analysis_log.txt
NO personal data (names / phones / emails) is written to output.
"""
import sys, io, re, datetime, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import pandas as pd
import numpy as np
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ============================================================
# PATHS
# ============================================================
CRM_FILE    = r"C:\Users\Sinaymer\Desktop\City volta solar\Residential - Last view used.xlsx"
SETTLE_CSV  = r"C:\Users\Sinaymer\Desktop\City volta solar\volta-settlements.csv"
OUTPUT_FILE = r"C:\Users\Sinaymer\Desktop\City volta solar\crm_knowledge_analysis_output.xlsx"
LOG_FILE    = r"C:\Users\Sinaymer\Desktop\City volta solar\crm_analysis_log.txt"

# ============================================================
# COLUMN MAP (index → internal name)
# ============================================================
COL_MAP = {
    4:   'Solar_Expert',
    10:  'Status',
    20:  'City',
    21:  'Regional_Council',
    26:  'Site_Description',
    28:  'Pre_Meeting_Remarks',
    29:  'Building_Type',
    33:  'Roof_Size',
    35:  'Roof_Type',
    37:  'False_SE_Call',
    38:  'False_SE_Call_Reasons',
    39:  'Irrelevant_Before',
    41:  'Date_Expert_Call',
    42:  'Date_Frontal_Meeting',
    43:  'Date_Video_Meeting',
    44:  'Date_Phone_Meeting',
    60:  'Lead_Source',
    78:  'Meeting_Summary',
    84:  'Irrelevant_After',
    87:  'Losing_Reason',
    130: 'CSL_Remarks',
    137: 'Installation_Notes',
    193: 'Electrical_Infra',
    251: 'Cancellation_Reason',
    299: 'Podio_ID',
    302: 'Internal_ID',
}

# ============================================================
# CONSTANTS — DECISION LOGIC
# ============================================================
ARAB_HE = {
    'אום אלפחם','שפרעם','טמרה','כפר כנא','כפר כנה','נצרת',
    'סחנין','עראבה','דיר חנא','ריינה','עיבלין','מגאר',
    "מג'ד אל-כרום",'כאבול','כפר יאסיף','אבו סנאן','ירכא',
    "ג'לג'וליה","ג'לג'ולייה",'קלנסווה','טייבה',
    'באקה אל-גרביה','באקה אלגרביה','כפר קרע','בועינה נוגידאת',
    'עין מאהל','כפר מנדא','עילוט','רהט','חורה','ערערה בנגב',
    'לקיה','כסיפה','תל שבע','ערערה',"ג'ת","ג'ת (כרמל)",
    'פסוטה',"ח'ורפיש",'חורפיש','נחף','עמקה',"יאנוח-ג'ת",
    "כפר סמיע","ג'ולס",'פוריידיס',"ג'סר א-זרקא",
    'כפר ברא','ברטעה','מעיליה','פקיעין',"ג'יס","עין קינייה",
    'עספיא','עסאפיא','מזרעה','בסמה','בענה','כסרא-סמיע',
    'ניין','שבלי - אום אלגנם','כפר עארה','כפר ברא',
    'עין ראפה','אבו גוש','בועינה','כפר יאנוח',
}
ARAB_EN = {
    "um al-fahm","umm el-fahm","shefa-'amr","shefa amr","shefaram","shafa-amr",
    "tamra","kafr kana","kfar kana","nazareth","sakhnin","arrabe","deir hanna",
    "reineh","eilabun","mghar","majd al-krum","kabul","kafr yasif","abu snan",
    "yarka","jaljulia","qalansawe","taibe","baka-jatt","baka al-gharbiyye",
    "kafr qara","bi'na","ein mahel","kafr manda","ilut","rahat","hura",
    "ar'ara banegev","lakiya","kseife","tel sheva","arara","jatt","fassouta",
    "hurfeish","nahaf","amka","yanuh-jat","kafr sumei","julis","fureidis",
    "jisr az-zarqa","kafr bara","bartaa","me'ilia","peki'in","jish","ain qinya",
    "isifya","ein rafa","kafr kara","beineh","bueineh",
}
SOUTH_HE = {
    'אילת','ספיר','פארן','שיזפון','יטבתה','לוטן','עין יהב',
    'חצבה','גרופית','שחרות','קטורה','סמר','נאות סמדר','ידן','עיידות','בקע',
}
SOUTH_EN = {
    'eilat','sapir','paran','shizafon','yotvata','lotan','ein yahav',
    'hatzeva','grofit','shaharut','ketura','samar','neot smadar','idan',
}
BAD_ROOF    = {'אסבסט','בנייה קלה','מייטק','בנייה קלה/מייטק'}
CAUTION_ROOF = {'פאנל מבודד','פנל מבודד','איסכורית','סנטף','סנטאף','פוליקרבונט'}
COMPLETED_ST = {'הפרויקט הושלם','הפרויקט הושלם - ללא תחזוקה','הפרויקט הושלם - עם תחזוקה'}

INFRA_TERMS = [
    ('אסבסט','אסבסט'),
    ('בנייה קלה','בנייה קלה'),('מייטק','מייטק'),
    ('פאנל מבודד','פאנל מבודד'),('פנל מבודד','פאנל מבודד'),
    ('איסכורית','איסכורית'),
    ('סנטף','סנטף'),('סנטאף','סנטף'),('פוליקרבונט','פוליקרבונט'),
    ('פרגולה סולארית','פרגולה סולארית'),
    ('פרגולה','פרגולה'),
    ('בטון שטוח','בטון שטוח'),('גג בטון','בטון שטוח'),
    ('גג רעפים','רעפים'),('רעפים','רעפים'),
]

# PII regex
EMAIL_RE  = re.compile(r'\[?[\w.+\-]+@[\w\-]+\.[a-zA-Z]{2,}\]?(?:\([^)]*\))?', re.I)
PHONE_RE  = re.compile(r'(?:\+972[\s\-.]?)?0[2-9][\d\s\-\.]{7,10}')
NAME_RE   = re.compile(r'^(?:Name|Phone|Email|City|Size):\s*.+$', re.M | re.I)

# Excel colours
CLR = {
    'header'   : 'FF1A3A2A',
    'מתקינים' : 'FFD8F3DC',
    'לא מתקינים': 'FFFFE8E8',
    'להתיייעץ': 'FFFFF3CD',
    'לא זוהה' : 'FFE8E8F0',
    'alt_row'  : 'FFF9FFFE',
}

# ============================================================
# HELPERS
# ============================================================
def v(val):
    s = str(val).strip() if pd.notna(val) and str(val).strip() not in ('nan','NaT','None','') else ''
    return s

def clean_pii(text):
    if not text:
        return ''
    t = NAME_RE.sub('', text)
    t = EMAIL_RE.sub('[מייל הוסר]', t)
    t = PHONE_RE.sub('[טל הוסר]', t)
    t = re.sub(r'\n{3,}', '\n\n', t).strip()
    return t[:300]

def norm_city(city_raw):
    city = v(city_raw)
    if not city:
        return ''
    # Remove ", Israel" suffix from English names
    city = re.sub(r',?\s*Israel$', '', city, flags=re.I).strip()
    return city

def is_arab(city):
    c = city.strip()
    if c in ARAB_HE:
        return True
    cl = c.lower()
    if cl in ARAB_EN:
        return True
    # Common Arabic patterns in Hebrew
    for pat in ["ג'","ח'","כ'"]:
        if pat in c and len(c) > 3:
            pass  # not conclusive alone
    return False

def is_south(city):
    c = city.strip()
    if c in SOUTH_HE:
        return True
    if c.lower() in SOUTH_EN:
        return True
    return False

def parse_roof_size(val):
    s = v(val)
    if not s:
        return None
    # "120+120" → 240
    if '+' in s:
        parts = re.findall(r'\d+', s)
        if parts:
            return sum(int(p) for p in parts)
    m = re.search(r'\d+', s)
    return int(m.group()) if m else None

def extract_infra(text):
    """Return list of infra types found in free text."""
    if not text:
        return []
    found = []
    seen = set()
    for term, canonical in INFRA_TERMS:
        if term in text and canonical not in seen:
            found.append(canonical)
            seen.add(canonical)
    return found

def passed_to_expert(row):
    if v(row['Solar_Expert']):
        return 'כן'
    if v(row['Date_Expert_Call']) or v(row['Date_Frontal_Meeting']) or v(row['Date_Video_Meeting']):
        return 'כן'
    if v(row['False_SE_Call']) == 'Yes':
        return 'לא (שיחה כוזבת)'
    if v(row['Irrelevant_Before']):
        return 'לא'
    return 'לא ידוע'

def decide(city, city_norm, status, roof_type_raw, roof_size_num, building_type,
           irrel_before, irrel_after, infra_list, existing_status, cancellation):
    """Returns (decision, reason, confidence, update_kb)"""
    if not city_norm:
        return 'לא זוהה', 'שם יישוב חסר', 'נמוכה', 'לבדיקה'
    if is_arab(city_norm):
        return 'לא מתקינים', 'יישוב ערבי', 'גבוהה', 'כן'
    if is_south(city_norm):
        return 'לא מתקינים', 'דרומית למצפה רמון', 'גבוהה', 'כן'
    if city_norm == 'מצפה רמון' or city_norm.lower() == 'mitzpe ramon':
        return 'מתקינים', 'מצפה רמון — חיובי לפי כלל', 'גבוהה', 'לבדיקה'

    # Completed project = confirmed install
    if any(s in status for s in COMPLETED_ST):
        return 'מתקינים', 'פרויקט הושלם', 'גבוהה', 'כן'

    # Combine roof info
    all_infra = set(infra_list)
    if roof_type_raw:
        all_infra.update(extract_infra(roof_type_raw))
        for t in roof_type_raw.split('+'):
            t = t.strip()
            if t:
                all_infra.add(t)

    # Asbestos → stop
    if 'אסבסט' in all_infra or 'אסבסט' in (roof_type_raw or ''):
        return 'לא מתקינים', 'אסבסט — לא לפני טיפול', 'גבוהה', 'לבדיקה'

    # Bad roof → stop
    for bad in BAD_ROOF:
        if bad in all_infra or bad in (roof_type_raw or ''):
            return 'לא מתקינים', f'סוג גג לא מתאים: {bad}', 'גבוהה', 'לבדיקה'

    # Shared condo
    if 'קונדו בבעלות משותפת' in (building_type or ''):
        return 'לא מתקינים', 'קונדו בבעלות משותפת', 'גבוהה', 'לא'

    # Roof too small
    if roof_size_num is not None and roof_size_num < 60:
        return 'לא מתקינים', f'שטח גג {roof_size_num} מ"ר — מתחת ל-60', 'גבוהה', 'לא'

    # Irrelevant - roof not suitable
    if 'גג לא מתאים' in (irrel_before or '') or 'גג לא מתאים' in (irrel_after or ''):
        return 'לא מתקינים', 'גג לא מתאים להתקנה (לפי CRM)', 'בינונית', 'לבדיקה'

    # Caution roof types
    for caut in CAUTION_ROOF:
        if caut in all_infra or caut in (roof_type_raw or ''):
            return 'להתיייעץ', f'סוג גג מצריך בדיקה: {caut}', 'בינונית', 'לבדיקה'

    # Private condo
    if 'קונדו בבעלות פרטית' in (building_type or ''):
        return 'להתיייעץ', 'קונדו בבעלות פרטית — נדרש אישור גישה לגג', 'בינונית', 'לא'

    # Borderline size
    if roof_size_num is not None and roof_size_num < 70:
        return 'להתיייעץ', f'שטח גג {roof_size_num} מ"ר — גבולי (60–69)', 'בינונית', 'לא'

    # Known settlement DB
    if existing_status == 'מתקינים':
        return 'מתקינים', 'מאגר יישובים קיים', 'בינונית', 'לא'
    if existing_status == 'לא מתקינים':
        return 'לא מתקינים', 'מאגר יישובים קיים', 'בינונית', 'לא'
    if existing_status == 'להתיייעץ':
        return 'להתיייעץ', 'מאגר יישובים קיים', 'בינונית', 'לא'

    # Irrelevant statuses without a clear reason
    if any(s in status for s in ['לא רלוונטי', 'irrelevant']):
        reason = irrel_before or irrel_after or cancellation or 'סיבה לא צוינה'
        return 'להתיייעץ', f'ליד לא רלוונטי — {reason}', 'נמוכה', 'לבדיקה'

    # Pergola only — not enough for installation
    if all_infra == {'פרגולה'} or all_infra == {'פרגולה סולארית'}:
        if not any(x in all_infra for x in ['בטון שטוח','רעפים']):
            return 'להתיייעץ', 'פרגולה בלבד — לא מבצעים לבד (צריך גג)', 'בינונית', 'לבדיקה'

    return 'להתיייעץ', 'אין מידע מספיק להחלטה', 'נמוכה', 'לבדיקה'

# ============================================================
# LOAD DATA
# ============================================================
print("טוען קובץ CRM...")
df_raw = pd.read_excel(CRM_FILE, sheet_name=0, header=0,
                        usecols=list(COL_MAP.keys()))
df_raw.columns = [COL_MAP[i] for i in COL_MAP.keys()]
total_rows = len(df_raw)
print(f"  נטענו {total_rows:,} שורות")

# Load settlement DB for cross-reference
settle_lookup = {}
if os.path.exists(SETTLE_CSV):
    df_s = pd.read_csv(SETTLE_CSV, encoding='utf-8-sig', header=0)
    for _, r in df_s.iterrows():
        name = v(r.iloc[0])
        status_s = v(r.iloc[2])
        if name:
            settle_lookup[name.strip()] = status_s
    print(f"  נטענו {len(settle_lookup):,} יישובים ממאגר הידע הקיים")

def lookup_settlement(city_norm):
    if city_norm in settle_lookup:
        return settle_lookup[city_norm]
    # Try partial match
    for k, v2 in settle_lookup.items():
        if city_norm and k and (city_norm in k or k in city_norm):
            return v2
    return ''

# ============================================================
# PROCESS ROWS
# ============================================================
print("מעבד שורות...")
decisions_rows  = []
open_q_rows     = []
city_decisions  = defaultdict(list)  # city → list of decisions
infra_cases     = defaultdict(list)  # infra_type → list of cases
rejection_reasons = []
top_cities_all  = []

rows_no_city    = 0
rows_arab       = 0
rows_south      = 0
rows_completed  = 0
rows_install    = 0
rows_no_install = 0
rows_consult    = 0
rows_unknown    = 0
rows_to_expert  = 0

for idx, row in df_raw.iterrows():
    row_num   = idx + 2  # +2 for header + 1-based
    city_raw  = v(row['City'])
    city_norm = norm_city(city_raw)

    if not city_norm:
        rows_no_city += 1

    top_cities_all.append(city_norm or '(ריק)')

    status        = v(row['Status'])
    roof_type_raw = v(row['Roof_Type'])
    roof_size_num = parse_roof_size(row['Roof_Size'])
    roof_size_disp= v(row['Roof_Size'])
    building_type = v(row['Building_Type'])
    irrel_before  = v(row['Irrelevant_Before'])
    irrel_after   = v(row['Irrelevant_After'])
    cancellation  = v(row['Cancellation_Reason'])
    site_desc_raw = v(row['Site_Description'])
    site_desc     = clean_pii(site_desc_raw)
    meeting_sum   = clean_pii(v(row['Meeting_Summary']))
    pre_remarks   = clean_pii(v(row['Pre_Meeting_Remarks']))
    install_notes = clean_pii(v(row['Installation_Notes']))
    csl_remarks   = clean_pii(v(row['CSL_Remarks']))
    losing_reason = v(row['Losing_Reason'])
    lead_source   = v(row['Lead_Source'])
    electrical    = v(row['Electrical_Infra'])
    podio_id      = v(row['Podio_ID'])
    internal_id   = v(row['Internal_ID'])

    # Infra from text
    infra_list = extract_infra(site_desc_raw + ' ' + v(row['Pre_Meeting_Remarks']) + ' ' + v(row['Meeting_Summary']))
    infra_display = ', '.join(infra_list) if infra_list else (roof_type_raw or '')

    # Passed to expert?
    expert_status = passed_to_expert(row)
    if expert_status == 'כן':
        rows_to_expert += 1

    # Existing DB status
    existing_status = lookup_settlement(city_norm)

    # Decision
    decision, reason, confidence, update_kb = decide(
        city_raw, city_norm, status, roof_type_raw, roof_size_num,
        building_type, irrel_before, irrel_after, infra_list, existing_status, cancellation
    )

    # Counters
    if decision == 'מתקינים':    rows_install    += 1
    elif decision == 'לא מתקינים': rows_no_install += 1
    elif decision == 'להתיייעץ':   rows_consult    += 1
    else:                          rows_unknown    += 1

    if is_arab(city_norm): rows_arab  += 1
    if is_south(city_norm): rows_south += 1
    if any(s in status for s in COMPLETED_ST): rows_completed += 1

    # Collect rejection reasons
    for r_src in [irrel_before, irrel_after, cancellation, losing_reason]:
        if r_src:
            rejection_reasons.append(r_src)

    # Note for agent — combine useful non-PII text
    agent_note_parts = []
    if site_desc:   agent_note_parts.append(site_desc[:150])
    if pre_remarks: agent_note_parts.append(pre_remarks[:100])
    if csl_remarks: agent_note_parts.append(csl_remarks[:100])
    agent_note = ' | '.join(agent_note_parts)[:300]

    record = {
        'מזהה רשומה'        : podio_id or internal_id or str(row_num),
        'שורה'              : row_num,
        'יישוב'             : city_norm,
        'מועצה אזורית'      : v(row['Regional_Council']),
        'סוג בניין'         : building_type,
        'סוג תשתית'         : infra_display or '—',
        'שטח גג'            : roof_size_disp or '—',
        'עבר למומחה'        : expert_status,
        'סטטוס CRM'         : status or '—',
        'החלטה מוצעת'       : decision,
        'סיבת החלטה'        : reason,
        'מקור החלטה'        : 'CRM',
        'רמת ודאות'         : confidence,
        'הערה לנציג'        : agent_note,
        'האם לעדכן מאגר'    : update_kb,
    }
    decisions_rows.append(record)

    # Track by city
    if city_norm:
        city_decisions[city_norm].append({
            'decision': decision, 'reason': reason,
            'status': status, 'infra': infra_display,
            'expert': expert_status, 'note': agent_note[:120],
            'row': row_num
        })

    # Track infra
    for infra_type in infra_list:
        infra_cases[infra_type].append({
            'city': city_norm, 'decision': decision,
            'roof_size': roof_size_disp, 'status': status, 'row': row_num
        })

    # Open questions
    if city_norm and not existing_status and decision == 'להתיייעץ' and confidence == 'נמוכה':
        open_q_rows.append({
            'נושא'          : f'יישוב לא מזוהה: {city_norm}',
            'תיאור'         : f'יישוב {city_norm!r} אינו במאגר הידע',
            'למה לא ברור'   : 'יישוב לא נמצא ב-volta-settlements.csv',
            'מה לשאול'      : f'האם {city_norm} נכלל בפריסה? מה הסטטוס?',
            'דוגמה CRM'     : (agent_note or status or '—')[:200],
            'עדיפות'        : 'בינונית',
        })
    if irrel_before == 'אחר - מפורט בהערות' and not agent_note:
        open_q_rows.append({
            'נושא'          : 'פסילה בלתי מוסברת',
            'תיאור'         : f'שורה {row_num}: "אחר - מפורט בהערות" ללא הערה',
            'למה לא ברור'   : 'הסיבה האמיתית לפסילה לא תועדה',
            'מה לשאול'      : 'מה הסיבה האמיתית לפסילה?',
            'דוגמה CRM'     : f'יישוב: {city_norm} | סטטוס: {status}',
            'עדיפות'        : 'נמוכה',
        })

print(f"  עובדו {len(decisions_rows):,} שורות")

# ============================================================
# BUILD SHEET 2 — SETTLEMENT UPDATES
# ============================================================
settle_upd_rows = []
for city, cases in city_decisions.items():
    if not city:
        continue
    # Settlement-level signals ONLY. Property-level rejections (small roof,
    # unsuitable roof type) are per-lead and must NOT override a city decision.
    completed_count = sum(1 for c in cases if c['status'] in COMPLETED_ST)
    has_completed   = completed_count > 0
    existing = lookup_settlement(city)

    if has_completed and existing != 'מתקינים':
        example = next((c['note'] for c in cases if c['status'] in COMPLETED_ST), '')
        settle_upd_rows.append({
            'יישוב'          : city,
            'סוג יישוב'      : '',
            'החלטה קיימת'    : existing or '—',
            'החלטה מוצעת'    : 'מתקינים',
            'סיבה'           : f'{completed_count} פרויקטים הושלמו בפועל ביישוב',
            'מקור'           : 'CRM',
            'דוגמה מהשיחה'   : (example or '')[:200],
            'רמת ודאות'      : 'גבוהה',
            'דורש אישור'     : 'לא',
        })
    elif is_arab(city) and existing != 'לא מתקינים':
        settle_upd_rows.append({
            'יישוב'          : city,
            'סוג יישוב'      : 'ערבי',
            'החלטה קיימת'    : existing or '—',
            'החלטה מוצעת'    : 'לא מתקינים',
            'סיבה'           : 'יישוב ערבי',
            'מקור'           : 'CRM',
            'דוגמה מהשיחה'   : '',
            'רמת ודאות'      : 'גבוהה',
            'דורש אישור'     : 'לא',
        })
    elif is_south(city) and existing != 'לא מתקינים' and city != 'מצפה רמון':
        settle_upd_rows.append({
            'יישוב'          : city,
            'סוג יישוב'      : '',
            'החלטה קיימת'    : existing or '—',
            'החלטה מוצעת'    : 'לא מתקינים',
            'סיבה'           : 'דרומית למצפה רמון',
            'מקור'           : 'CRM',
            'דוגמה מהשיחה'   : '',
            'רמת ודאות'      : 'גבוהה',
            'דורש אישור'     : 'לא',
        })

# ============================================================
# BUILD SHEET 3 — INFRASTRUCTURE UPDATES
# ============================================================
infra_upd_rows = []
infra_decisions_map = {
    'בטון שטוח'        : ('מתקינים', 'גג בטון שטוח — מותאם לרוב', 'ודא טופס 4 וגודל', 'גבוהה'),
    'רעפים'            : ('מתקינים', 'גג רעפים — בדוק גיל', 'רעפים מעל 25 שנה — להתיייעץ', 'גבוהה'),
    'פרגולה סולארית'   : ('להתיייעץ', 'פרגולה סולארית — פאנלים מיוחדים', 'רק עם אישור מומחה', 'בינונית'),
    'פרגולה'           : ('להתיייעץ', 'פרגולה בלבד — לא מבצעים לבד', 'חייב להיות גג ראשי בנוסף', 'בינונית'),
    'פאנל מבודד'       : ('להתיייעץ', 'פאנל מבודד — נדרש אישור מנהל', 'יש מקרים שאושרו בעבר', 'בינונית'),
    'איסכורית'         : ('להתיייעץ', 'איסכורית — נדרש קונסטרוקטור', 'אישור מבני לפני כל התקדמות', 'בינונית'),
    'סנטף'             : ('להתיייעץ', 'סנטף — לא לקבוע לבד', 'לבדוק עם מנהל', 'נמוכה'),
    'פוליקרבונט'       : ('להתיייעץ', 'פוליקרבונט — לא מתאים בד"כ', 'לבדוק עם מנהל', 'נמוכה'),
    'אסבסט'            : ('לא מתקינים', 'אסבסט — לא מתקינים', 'חייב טיפול מוקדם של אסבסט', 'גבוהה'),
    'בנייה קלה'        : ('לא מתקינים', 'בנייה קלה — לא מתקינים', '', 'גבוהה'),
    'מייטק'            : ('לא מתקינים', 'מייטק — לא מתקינים', '', 'גבוהה'),
}

for infra_type, cases in infra_cases.items():
    count = len(cases)
    completed = sum(1 for c in cases if c['status'] in COMPLETED_ST)
    example_cities = list({c['city'] for c in cases[:3]})
    example_str = ', '.join(example_cities)
    defaults = infra_decisions_map.get(infra_type, ('להתיייעץ', infra_type, '', 'נמוכה'))
    infra_upd_rows.append({
        'סוג תשתית'      : infra_type,
        'כמות מקרים'     : count,
        'פרויקטים הושלמו': completed,
        'ערים לדוגמה'    : example_str,
        'החלטה מוצעת'    : defaults[0],
        'תיאור'          : defaults[1],
        'תנאים'          : defaults[2],
        'מקור'           : 'CRM + כללים קיימים',
        'רמת ודאות'      : defaults[3],
        'הערה לנציג'     : defaults[2] or '—',
    })

# ============================================================
# BUILD SHEET 5 — SUMMARY
# ============================================================
top_cities = Counter(c for c in top_cities_all if c and c != '(ריק)').most_common(10)
top_rejections = Counter(rejection_reasons).most_common(10)
top_statuses   = Counter(v(r) for r in df_raw['Status'] if v(r)).most_common(10)

# ============================================================
# WRITE EXCEL
# ============================================================
print("כותב קובץ Excel...")
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

HEADER_FONT  = Font(name='Arial', bold=True, color='FFFFFFFF', size=11)
HEADER_FILL  = PatternFill('solid', fgColor=CLR['header'])
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN   = Alignment(horizontal='right', vertical='top', wrap_text=True)
ALT_FILL     = PatternFill('solid', fgColor=CLR['alt_row'])
DECISION_FILLS = {
    'מתקינים'     : PatternFill('solid', fgColor=CLR['מתקינים']),
    'לא מתקינים'  : PatternFill('solid', fgColor=CLR['לא מתקינים']),
    'להתיייעץ'    : PatternFill('solid', fgColor=CLR['להתיייעץ']),
    'לא זוהה'     : PatternFill('solid', fgColor=CLR['לא זוהה']),
}

def add_sheet(wb, name, df, decision_col=None, col_widths=None):
    ws = wb.create_sheet(title=name)
    ws.sheet_view.rightToLeft = True

    headers = list(df.columns)
    for col_i, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=hdr)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        ws.row_dimensions[1].height = 30

    for row_i, record in enumerate(df.itertuples(index=False), 2):
        for col_i, val in enumerate(record, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val if val != '—' else val)
            cell.alignment = CELL_ALIGN
            # Alternate row shading
            if row_i % 2 == 0:
                cell.fill = ALT_FILL
        # Decision colour
        if decision_col and decision_col in headers:
            d_idx = headers.index(decision_col) + 1
            d_val = str(record[headers.index(decision_col)])
            if d_val in DECISION_FILLS:
                for c in range(1, len(headers)+1):
                    ws.cell(row=row_i, column=c).fill = DECISION_FILLS[d_val]

    # Column widths
    if col_widths:
        for col_i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_i)].width = width
    else:
        for col_i in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col_i)].width = 18

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    return ws

# ----- Sheet 1: Decisions -----
df_dec = pd.DataFrame(decisions_rows)
add_sheet(wb, 'ניתוח CRM - החלטות', df_dec, decision_col='החלטה מוצעת',
          col_widths=[14,7,18,16,14,20,10,12,18,16,30,12,12,35,12])

# ----- Sheet 2: Settlement updates -----
if settle_upd_rows:
    df_su = pd.DataFrame(settle_upd_rows)
else:
    df_su = pd.DataFrame(columns=['יישוב','סוג יישוב','החלטה קיימת','החלטה מוצעת',
                                   'סיבה','מקור','דוגמה מהשיחה','רמת ודאות','דורש אישור'])
add_sheet(wb, 'עדכונים מוצעים - יישובים', df_su, decision_col='החלטה מוצעת',
          col_widths=[18,12,14,14,28,10,40,12,12])

# ----- Sheet 3: Infra updates -----
if infra_upd_rows:
    df_iu = pd.DataFrame(infra_upd_rows)
else:
    df_iu = pd.DataFrame(columns=['סוג תשתית','כמות מקרים','פרויקטים הושלמו','ערים לדוגמה',
                                   'החלטה מוצעת','תיאור','תנאים','מקור','רמת ודאות','הערה לנציג'])
add_sheet(wb, 'עדכונים מוצעים - תשתיות', df_iu, decision_col='החלטה מוצעת',
          col_widths=[16,12,14,30,14,30,30,14,12,30])

# ----- Sheet 4: Open questions -----
df_oq = pd.DataFrame(open_q_rows) if open_q_rows else pd.DataFrame(
    columns=['נושא','תיאור','למה לא ברור','מה לשאול','דוגמה CRM','עדיפות'])
add_sheet(wb, 'שאלות פתוחות', df_oq, col_widths=[30,35,30,35,40,12])

# ----- Sheet 5: Summary -----
ws_sum = wb.create_sheet(title='סיכום')
ws_sum.sheet_view.rightToLeft = True
ws_sum.column_dimensions['A'].width = 35
ws_sum.column_dimensions['B'].width = 18

def sum_row(ws, row, label, value, bold=False):
    a = ws.cell(row=row, column=1, value=label)
    b = ws.cell(row=row, column=2, value=value)
    a.alignment = Alignment(horizontal='right')
    b.alignment = Alignment(horizontal='center')
    if bold:
        a.font = Font(bold=True, size=12)
        b.font = Font(bold=True, size=12)
    return row + 1

r = 1
ws_sum.cell(row=r, column=1, value='סיכום ניתוח CRM — וולטה סולאר').font = Font(bold=True, size=14)
ws_sum.merge_cells('A1:B1')
r = 2
r = sum_row(ws_sum, r, '─────────── נתונים כלליים ───────────', '', bold=True)
r = sum_row(ws_sum, r, 'סה"כ שורות בקובץ CRM',   f'{total_rows:,}')
r = sum_row(ws_sum, r, 'שורות ללא שם יישוב',       f'{rows_no_city:,}')
r = sum_row(ws_sum, r, 'שורות שעובדו',             f'{len(decisions_rows):,}')
r = sum_row(ws_sum, r, 'מזוהות כעברו למומחה',      f'{rows_to_expert:,}')
r = sum_row(ws_sum, r, 'פרויקטים שהושלמו',         f'{rows_completed:,}')
r += 1
r = sum_row(ws_sum, r, '─────────── החלטות מוצעות ───────────', '', bold=True)
r = sum_row(ws_sum, r, 'מתקינים',                  f'{rows_install:,}')
r = sum_row(ws_sum, r, 'לא מתקינים',               f'{rows_no_install:,}')
r = sum_row(ws_sum, r, 'להתיייעץ',                 f'{rows_consult:,}')
r = sum_row(ws_sum, r, 'לא זוהה / חסר מידע',       f'{rows_unknown:,}')
r = sum_row(ws_sum, r, 'מתוכם יישובים ערביים',     f'{rows_arab:,}')
r = sum_row(ws_sum, r, 'מתוכם דרומי למצפה רמון',   f'{rows_south:,}')
r += 1
r = sum_row(ws_sum, r, '─────────── 10 יישובים נפוצים ────────', '', bold=True)
for city_name, cnt in top_cities:
    r = sum_row(ws_sum, r, city_name, f'{cnt} רשומות')
r += 1
r = sum_row(ws_sum, r, '─────────── 10 סיבות פסילה נפוצות ───', '', bold=True)
for reason_name, cnt in top_rejections:
    r = sum_row(ws_sum, r, reason_name[:50], f'{cnt} פעמים')
r += 1
r = sum_row(ws_sum, r, '─────────── 10 סטטוסים נפוצים ────────', '', bold=True)
for st_name, cnt in top_statuses:
    r = sum_row(ws_sum, r, st_name[:50], f'{cnt} פעמים')

wb.save(OUTPUT_FILE)
print(f"  ✓ נשמר: {OUTPUT_FILE}")

# ============================================================
# WRITE LOG
# ============================================================
log_lines = [
    f"=== לוג ניתוח CRM — וולטה סולאר ===",
    f"תאריך ריצה: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
    f"קובץ מקור:  {CRM_FILE}",
    f"קובץ פלט:   {OUTPUT_FILE}",
    f"",
    f"--- נתונים כלליים ---",
    f"סה\"כ שורות בקובץ:      {total_rows:,}",
    f"שורות ללא יישוב:       {rows_no_city:,}",
    f"שורות שעובדו:          {len(decisions_rows):,}",
    f"",
    f"--- עמודות שזוהו ---",
] + [f"  col[{k:3d}] = {v2}" for k, v2 in COL_MAP.items()] + [
    f"",
    f"--- החלטות ---",
    f"מתקינים:               {rows_install:,}",
    f"לא מתקינים:            {rows_no_install:,}",
    f"להתיייעץ:              {rows_consult:,}",
    f"לא זוהה:               {rows_unknown:,}",
    f"עברו למומחה:            {rows_to_expert:,}",
    f"פרויקטים שהושלמו:      {rows_completed:,}",
    f"יישובים ערביים:        {rows_arab:,}",
    f"דרומי למצפה רמון:      {rows_south:,}",
    f"",
    f"--- עדכונים מוצעים ---",
    f"יישובים לעדכון:        {len(settle_upd_rows):,}",
    f"סוגי תשתית שנמצאו:    {len(infra_cases):,}",
    f"שאלות פתוחות:          {len(open_q_rows):,}",
    f"",
    f"--- יישובים נפוצים (10 ראשונים) ---",
] + [f"  {n}: {c}" for n, c in top_cities] + [
    f"",
    f"--- סיבות פסילה נפוצות (10 ראשונות) ---",
] + [f"  {n}: {c}" for n, c in top_rejections]

with open(LOG_FILE, 'w', encoding='utf-8') as f:
    f.write('\n'.join(log_lines))
print(f"  ✓ לוג נשמר: {LOG_FILE}")
print("\nהניתוח הסתיים בהצלחה!")
